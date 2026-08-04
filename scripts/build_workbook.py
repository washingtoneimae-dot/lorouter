"""
Build performance_data.xlsx from the canonical figures in TECHNICAL.md
(kept in sync with the document's numbers by construction). Uses Excel
formulas for every derived column (percent changes, aggregates) rather than
pasting computed results, per xlsx skill guidance.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

FONT = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=13)
NOTE_FONT = Font(name=FONT, italic=True, size=9, color="595959")
BODY_FONT = Font(name=FONT, size=10)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_bar_chart(ws, title, y_title, cat_ref, data_ref, anchor, height=8, width=16, data_labels=False):
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = None
    chart.height = height
    chart.width = width
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.style = 10
    if data_labels:
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
    ws.add_chart(chart, anchor)
    return chart



def title_block(ws, title, note, ncols):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A2"] = note
    ws["A2"].font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.row_dimensions[2].height = 30
    for row in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")


# ============================================================
# Sheet 1: Summary
# ============================================================
ws = wb.create_sheet("Summary")
title_block(ws, "Profile-Routed MoE -- Performance Data Summary",
            "Fresh canonical run, all 6 reproduction scripts. See TECHNICAL.md for full context, formulas, and proofs. "
            "Exact figures vary run-to-run for multi-stage pipelines (see TECHNICAL.md section 12) -- qualitative "
            "findings are the reproducible claim.", 4)
headers = ["Script", "Sheet(s)", "Headline finding", "Reproduces exactly every run?"]
ws.append([])
ws.append(headers)
style_header_row(ws, 4, 4)
rows = [
    ["addition_isolation_suite.py", "Swap Isolation, Addition Flips, MSE by Domain, Multi-Seed Stability",
     "Swap isolation: 0.00% collateral change on untouched domains. Addition: 4 flips found, gating fix restores 3/4.", "No -- pattern reproduces, exact figures vary"],
    ["capacity_ablation.py", "Capacity Ablation",
     "326x more gate parameters: no improvement on genuinely ambiguous cases (63.6% vs 62.5% accuracy). Bayes-error, not capacity-bound.", "Yes -- exact to the decimal"],
    ["multi_dimension_compounding.py", "Multi-Dimension Compounding",
     "3 simultaneous domain additions at 1% FPR each compound to 3.00% aggregate; Bonferroni correction restores 1.01%.", "Yes -- compounding % exact; recall figures vary slightly"],
    ["text_validation.py", "Real Text Validation, Printer Prototype",
     "Addition-flip mechanism replicates on real text (6 flips, concentrated on boundary prompts). Systematic calibration-data generation fixes a degenerate threshold.", "Part 1 varies; Part 2 exact to the decimal"],
    ["boundary_solutions.py", "Boundary Solutions Comparison",
     "Of 4 mitigations tested, only the gated one-vs-rest fix (D) changes which expert wins top-1; fixes 2 of 3 flips.", "No -- pattern reproduces, exact figures vary"],
    ["build_workbook.py", "(generates this workbook)",
     "Regenerates performance_data.xlsx from the canonical figures; every derived column is a live Excel formula.", "Yes -- deterministic"],
]
for r in rows:
    ws.append(r)
for row in ws.iter_rows(min_row=5, max_row=4+len(rows)):
    for cell in row:
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
autosize(ws, [26, 34, 55, 28])
ws.freeze_panes = "A5"

# ============================================================
# Sheet 2: Swap Isolation
# ============================================================
ws = wb.create_sheet("Swap Isolation")
title_block(ws, "Swap Isolation (addition_isolation_suite.py, section_2_swap_isolation)",
            "Expert_code is swapped for a deliberately different (undertrained) alternative. Domains not swapped should show ~0% change.", 4)
ws.append([])
ws.append(["Domain", "MSE before", "MSE after", "% change"])
style_header_row(ws, 4, 4)
data = [
    ["code", 0.19995, 1.78403],
    ["creative", 0.03786, 0.03786],
    ["math", 0.08639, 0.08639],
    ["reasoning", 0.01046, 0.01046],
]
start = 5
for i, (name, before, after) in enumerate(data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    ws.cell(row=r, column=2, value=before).font = BODY_FONT
    ws.cell(row=r, column=3, value=after).font = BODY_FONT
    f = ws.cell(row=r, column=4, value=f"=(C{r}-B{r})/B{r}")
    f.font = BODY_FONT
    f.number_format = "+0.00%;-0.00%"
    ws.cell(row=r, column=2).number_format = "0.00000"
    ws.cell(row=r, column=3).number_format = "0.00000"
note_row = start + len(data) + 1
ws.cell(row=note_row, column=1, value="Only the swapped domain (code) changes materially. The other three show exactly 0.00% -- the isolation property.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
autosize(ws, [14, 14, 14, 12])
ws.freeze_panes = "A5"
add_bar_chart(ws, "MSE Before vs After Swap, by Domain", "MSE",
              Reference(ws, min_col=1, min_row=5, max_row=8),
              Reference(ws, min_col=2, max_col=3, min_row=4, max_row=8),
              "F4")

# ============================================================
# Sheet 3: Addition Flips & Gating Fix
# ============================================================
ws = wb.create_sheet("Addition Flips & Gating Fix")
title_block(ws, "Addition Flips + Gated Fix (addition_isolation_suite.py, section_3_addition_and_gating)",
            "Gate threshold calibrated at the 99th percentile on 8,000 held-out base-domain points. 'Gated' = law-detector score cleared the threshold.", 5)
ws.append([])
ws.append(["Sample (domain#id)", "Law detector score", "Gated?", "Result after gating", "Fixed?"])
style_header_row(ws, 4, 5)
flip_data = [
    ["creative#134", 0.701, "no", "Expert_creative (correct)", "FIXED"],
    ["math#34", 0.640, "no", "Expert_math (correct)", "FIXED"],
    ["math#36", 0.441, "no", "Expert_math (correct)", "FIXED"],
    ["math#108", 0.794, "yes", "Expert_law (incorrect)", "still wrong"],
]
start = 5
for i, row_ in enumerate(flip_data):
    r = start + i
    for c, val in enumerate(row_, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        if c == 5 and val == "FIXED":
            cell.font = Font(name=FONT, size=10, color="006100", bold=True)
        elif c == 5:
            cell.font = Font(name=FONT, size=10, color="9C0006", bold=True)
ws.cell(row=start+4, column=1, value="Fixed").font = Font(name=FONT, bold=True)
ws.cell(row=start+4, column=2, value=f"=COUNTIF(E{start}:E{start+3},\"FIXED\")&\"/\"&COUNTA(E{start}:E{start+3})").font = BODY_FONT
ws.cell(row=start+5, column=1, value="Genuine law recall at this threshold").font = Font(name=FONT, bold=True)
ws.cell(row=start+5, column=2, value=0.887).font = BODY_FONT
ws.cell(row=start+5, column=2).number_format = "0.0%"
note_row = start + 7
ws.cell(row=note_row, column=1, value="math#108's score (0.794) clears the calibrated threshold -- an honest, well-calibrated detector genuinely cannot separate it from a real law input without sacrificing recall elsewhere (see TECHNICAL.md 4.5 corollary).").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
autosize(ws, [20, 16, 10, 24, 14])
ws.freeze_panes = "A5"

# ============================================================
# Sheet 4: MSE by Domain (Addition Impact)
# ============================================================
ws = wb.create_sheet("MSE by Domain (Addition Impact)")
title_block(ws, "Aggregate MSE by Domain, Before vs After Addition",
            "Frozen v1 (safe) routing vs jointly-retrained v2 (broken) routing, across each domain's full test set. Generated fresh by this script -- no external file needed.", 4)
ws.append([])
ws.append(["Domain", "MSE before (v1, frozen)", "MSE after (v2, jointly-retrained)", "% change"])
style_header_row(ws, 4, 4)
mse_data = [
    ["code", 0.76979, 0.76979],
    ["creative", 0.03786, 0.06567],
    ["math", 0.08639, 0.26509],
    ["reasoning", 0.48072, 0.48072],
]
start = 5
for i, (name, before, after) in enumerate(mse_data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    ws.cell(row=r, column=2, value=before).font = BODY_FONT
    ws.cell(row=r, column=2).number_format = "0.00000"
    ws.cell(row=r, column=3, value=after).font = BODY_FONT
    ws.cell(row=r, column=3).number_format = "0.00000"
    f = ws.cell(row=r, column=4, value=f"=(C{r}-B{r})/B{r}")
    f.font = BODY_FONT
    f.number_format = "+0.00%;-0.00%"
note_row = start + len(mse_data) + 1
ws.cell(row=note_row, column=1, value="Two of four untouched domains show real collateral damage (creative, math); the other two show none in this run -- the failure concentrates on specific poorly-resolved regions, not uniform degradation.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
autosize(ws, [14, 22, 26, 12])
ws.freeze_panes = "A5"
add_bar_chart(ws, "MSE Before vs After Adding a New Domain, by Domain", "MSE",
              Reference(ws, min_col=1, min_row=5, max_row=8),
              Reference(ws, min_col=2, max_col=3, min_row=4, max_row=8),
              "F4")

# ============================================================
# Sheet 5: Multi-Seed Stability
# ============================================================
ws = wb.create_sheet("Multi-Seed Stability")
title_block(ws, "Flip Stability Across 5 Independent Training Seeds",
            "Same data distributions, 5 different model-initialization seeds. A stable core recurs regardless of seed; a shifting penumbra does not.", 3)
ws.append([])
ws.append(["Flip point (domain#id)", "Seeds it appears in (of 5)", "Category"])
style_header_row(ws, 4, 3)
stability_data = [
    ["math#108", 5, "stable core"],
    ["math#34", 5, "stable core"],
    ["code#115", 5, "stable core"],
    ["creative#134", 5, "stable core"],
    ["reasoning#146", 5, "stable core"],
    ["math#36", 2, "shifting penumbra"],
    ["reasoning#149", 2, "shifting penumbra"],
    ["code#82", 2, "shifting penumbra"],
    ["code#70", 1, "shifting penumbra"],
]
start = 5
for i, (name, count, cat) in enumerate(stability_data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    c2 = ws.cell(row=r, column=2, value=count)
    c2.font = BODY_FONT
    c2.alignment = Alignment(horizontal="center")
    c3 = ws.cell(row=r, column=3, value=cat)
    c3.font = Font(name=FONT, size=10, bold=(cat == "stable core"),
                    color="9C0006" if cat == "stable core" else "595959")
note_row = start + len(stability_data) + 1
ws.cell(row=note_row, column=1, value="5 of 9 unique flip points recur in every single seed -- tied to a specific, consistently poorly-resolved region, not training noise.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
autosize(ws, [24, 24, 20])
ws.freeze_panes = "A5"
add_bar_chart(ws, "Flip Frequency Across 5 Seeds", "Seeds (of 5)",
              Reference(ws, min_col=1, min_row=start, max_row=start+len(stability_data)-1),
              Reference(ws, min_col=2, min_row=4, max_row=start+len(stability_data)-1),
              "E4", data_labels=True)

# ============================================================
# Sheet 6: Capacity Ablation
# ============================================================
ws = wb.create_sheet("Capacity Ablation")
title_block(ws, "Does More Gate Capacity Resolve Genuinely Ambiguous Cases? (capacity_ablation.py)",
            "33-parameter gate vs 10,753-parameter gate (326x), identical training data. Reproduces exactly, every run.", 3)
ws.append([])
ws.append(["Metric", "Small gate (33 params)", "Large gate (10,753 params)"])
style_header_row(ws, 4, 3)
cap_data = [
    ["Overall AUC", 0.99604, 0.99614],
    ["Accuracy on ambiguous subset (2.53% of test set)", 0.636, 0.625],
    ["Overall accuracy", 0.97610, 0.97570],
]
start = 5
for i, (name, s, l) in enumerate(cap_data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    ws.cell(row=r, column=2, value=s).font = BODY_FONT
    ws.cell(row=r, column=2).number_format = "0.00000" if "AUC" in name else "0.000%"
    ws.cell(row=r, column=3, value=l).font = BODY_FONT
    ws.cell(row=r, column=3).number_format = "0.00000" if "AUC" in name else "0.000%"
r = start + len(cap_data)
ws.cell(row=r, column=1, value="Cases: large fixes small's error").font = BODY_FONT
ws.cell(row=r, column=2, value=32).font = BODY_FONT
r += 1
ws.cell(row=r, column=1, value="Cases: small fixes large's error").font = BODY_FONT
ws.cell(row=r, column=2, value=36).font = BODY_FONT
r += 1
ws.cell(row=r, column=1, value="Ratio (near 1.0 = no systematic advantage)").font = Font(name=FONT, bold=True)
ws.cell(row=r, column=2, value=f"=B{r-2}/B{r-1}").font = Font(name=FONT, bold=True)
ws.cell(row=r, column=2).number_format = "0.00"
note_row = r + 2
ws.cell(row=note_row, column=1, value="326x more parameters buys no measurable improvement on the hard cases -- the ambiguity is a Bayes-error property of the data, not a capacity bottleneck.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
autosize(ws, [42, 22, 24])
ws.freeze_panes = "A5"
add_bar_chart(ws, "Small vs Large Gate: No Systematic Advantage", "Score",
              Reference(ws, min_col=1, min_row=5, max_row=7),
              Reference(ws, min_col=2, max_col=3, min_row=4, max_row=7),
              "E4")

# ============================================================
# Sheet 7: Multi-Dimension Compounding
# ============================================================
ws = wb.create_sheet("Multi-Dimension Compounding")
title_block(ws, "Adding 3 Domains Simultaneously (multi_dimension_compounding.py)",
            "Each gate independently calibrated to 1% FPR on old-domain data. Compounding % reproduces exactly every run; recall figures vary slightly.", 4)
ws.append([])
ws.append(["Field", "Individual FPR (uncorrected)", "Recall (uncorrected)", "Recall (Bonferroni-corrected)"])
style_header_row(ws, 4, 4)
comp_data = [
    ["finance", 0.01, 0.700, 0.507],
    ["law", 0.01, 0.880, 0.800],
    ["medicine", 0.01, 0.613, 0.447],
]
start = 5
for i, (name, fpr, r_unc, r_cor) in enumerate(comp_data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    for c, v in [(2, fpr), (3, r_unc), (4, r_cor)]:
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = BODY_FONT
        cell.number_format = "0.0%"
r = start + len(comp_data) + 1
labels_vals = [
    ("Aggregate FPR, uncorrected (>=1 gate fires)", 0.0300),
    ("Naive independence prediction (1-(1-0.01)^3)", "=1-(1-0.01)^3"),
    ("Aggregate FPR, Bonferroni-corrected", 0.0101),
    ("Target aggregate FPR", 0.01),
]
for label, val in labels_vals:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True)
    cell = ws.cell(row=r, column=2, value=val)
    cell.font = Font(name=FONT, bold=True)
    cell.number_format = "0.00%"
    r += 1
note_row = r + 1
ws.cell(row=note_row, column=1, value="3 independent gates at 1% each compound to 3.00% aggregate (matches the naive independence prediction almost exactly) -- a multiple-comparisons problem, not a calibration bug. Bonferroni correction restores the target but costs real recall.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
autosize(ws, [40, 24, 20, 24])
ws.freeze_panes = "A5"
add_bar_chart(ws, "Recall by Field: Uncorrected vs Bonferroni-Corrected", "Recall",
              Reference(ws, min_col=1, min_row=5, max_row=7),
              Reference(ws, min_col=3, max_col=4, min_row=4, max_row=7),
              "F4")
add_bar_chart(ws, "Aggregate False-Capture Rate: Compounding vs Correction", "Rate",
              Reference(ws, min_col=1, min_row=9, max_row=12),
              Reference(ws, min_col=2, min_row=8, max_row=12),
              "F20")

# ============================================================
# Sheet 8: Real Text Validation
# ============================================================
ws = wb.create_sheet("Real Text Validation")
title_block(ws, "Does the Flip Mechanism Replicate on Real Text? (text_validation.py, part 1)",
            "TF-IDF+SVD embeddings (lexical, not full semantic -- no network access to a trained embedding model). 131 base-domain test prompts, 14 deliberately cross-domain.", 4)
ws.append([])
ws.append(["True domain", "Base profiler picked", "Joint (broken) profiler picked", "Prompt (truncated)"])
style_header_row(ws, 4, 4)
text_flips = [
    ["code", "math", "code", "draft terms of service clauses governing API usage limits for this app"],
    ["code", "math", "code", "assess whether this code snippet infringes on a competitor's registered..."],
    ["math", "reasoning", "law", "calculate the statute of limitations deadline given the filing date and..."],
    ["math", "creative", "law", "derive the amortization schedule required under the loan agreement's..."],
    ["creative", "creative", "law", "compose a courtroom drama scene where the plaintiff's testimony..."],
    ["reasoning", "creative", "law", "assess if the contract's ambiguous clause creates a valid basis for..."],
]
start = 5
for i, row_ in enumerate(text_flips):
    r = start + i
    for c, val in enumerate(row_, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True) if c == 4 else Alignment()
note_row = start + len(text_flips) + 1
ws.cell(row=note_row, column=1, value="6 flips, 100% on prompts that are genuinely cross-domain by construction; 0% on the 117 ordinary prompts. Every flip listed here involves law-adjacent vocabulary in a code/math/creative/reasoning prompt.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
autosize(ws, [14, 20, 26, 55])
ws.freeze_panes = "A5"

# ============================================================
# Sheet 9: Printer Prototype (Calibration)
# ============================================================
ws = wb.create_sheet("Printer Prototype (Calibration)")
title_block(ws, "Systematic vs Hand-Picked Boundary Calibration (text_validation.py, part 2)",
            "Reproduces exactly, every run. Compares calibrating a gate threshold on clean-only data vs clean+systematically-generated boundary data.", 3)
ws.append([])
ws.append(["Metric", "Clean-only calibration", "Printed (clean + systematic boundary)"])
style_header_row(ws, 4, 3)
printer_data = [
    ["Calibrated threshold", 0.0031, 0.9943],
    ["Genuine law recall", 1.000, 0.920],
    ["False-capture on fresh unambiguous data", 0.0160, 0.0000],
]
start = 5
for i, (name, a, b) in enumerate(printer_data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    ws.cell(row=r, column=2, value=a).font = BODY_FONT
    ws.cell(row=r, column=3, value=b).font = BODY_FONT
    if "threshold" in name.lower():
        ws.cell(row=r, column=2).number_format = "0.0000"
        ws.cell(row=r, column=3).number_format = "0.0000"
    else:
        ws.cell(row=r, column=2).number_format = "0.00%"
        ws.cell(row=r, column=3).number_format = "0.00%"
note_row = start + len(printer_data) + 1
ws.cell(row=note_row, column=1, value="300 hand-written calibration sentences produced a degenerate threshold (fires on almost everything). ~600 systematically-generated boundary examples fixed it, at a real cost: 8 points of law recall.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
autosize(ws, [42, 24, 34])
ws.freeze_panes = "A5"
add_bar_chart(ws, "Clean-Only vs Systematic Calibration", "Value (0-1 scale)",
              Reference(ws, min_col=1, min_row=5, max_row=7),
              Reference(ws, min_col=2, max_col=3, min_row=4, max_row=7),
              "F4")

# ============================================================
# Sheet 10: Boundary Solutions Comparison
# ============================================================
ws = wb.create_sheet("Boundary Solutions Comparison")
title_block(ws, "Four Mitigations, Head-to-Head (boundary_solutions.py)",
            "A/B/C are the original mitigations; D (gated one-vs-rest) is new, added this review pass. Same 3 boundary samples, same harness.", 5)
ws.append([])
ws.append(["Solution", "Mean error", "% vs baseline", "Samples improved (of 3)", "Changes WHICH expert wins top-1?"])
style_header_row(ws, 4, 5)
sol_data = [
    ["Baseline (broken)", 2.1466069276, None, None, "--"],
    ["A: Local Confidence", 2.1466069244, -0.001, 2, "No"],
    ["B: Adaptive tau", 1.4536902334, -0.323, 3, "No"],
    ["C: Variance Penalty", 2.1466069292, 0.0002, 1, "No"],
    ["D: Gated One-vs-Rest", 1.3046079307, -0.392, 2, "Yes -- fixes 2 of 3"],
]
start = 5
baseline_row = start
for i, (name, err, pct, imp, changes) in enumerate(sol_data):
    r = start + i
    ws.cell(row=r, column=1, value=name).font = BODY_FONT
    ws.cell(row=r, column=2, value=err).font = BODY_FONT
    ws.cell(row=r, column=2).number_format = "0.0000"
    if pct is not None:
        f = ws.cell(row=r, column=3, value=f"=(B{r}-B{baseline_row})/B{baseline_row}")
        f.font = BODY_FONT
        f.number_format = "+0.0%;-0.0%"
    else:
        ws.cell(row=r, column=3, value="--").font = BODY_FONT
    ws.cell(row=r, column=4, value=imp if imp is not None else "--").font = BODY_FONT
    cell5 = ws.cell(row=r, column=5, value=changes)
    cell5.font = Font(name=FONT, size=10, bold=changes.startswith("Yes"), color="006100" if changes.startswith("Yes") else "595959")
note_row = start + len(sol_data) + 1
ws.cell(row=note_row, column=1, value="Only D operates before the broken profile is computed (frozen base + independently-calibrated gate), so it's the only one that can re-select the correct expert rather than just reweight around the wrong one.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
autosize(ws, [24, 14, 16, 22, 30])
ws.freeze_panes = "A5"
add_bar_chart(ws, "Mean Error by Solution", "Mean error",
              Reference(ws, min_col=1, min_row=start, max_row=start+len(sol_data)-1),
              Reference(ws, min_col=2, min_row=4, max_row=start+len(sol_data)-1),
              "G4", data_labels=True)

# ============================================================
# Sheet 11: New Profiler Dataset Design (Draft)
# ============================================================
ws = wb.create_sheet("New Profiler Dataset Design")
title_block(ws, "Draft Dataset Design for Training a New Profiler",
            "Design, not measured data -- a template showing how a calibration dataset should be structured, based on findings elsewhere in this workbook: "
            "boundary examples are load-bearing (Printer Prototype sheet), calibration must be a held-out split separate from training (TECHNICAL.md 4.3), "
            "and per-domain worst-case calibration is safer than pooled aggregate when domains will later be composed (TECHNICAL.md section 9).", 9)
ws.append([])

# --- Schema definition ---
ws.append(["SCHEMA"])
ws["A6"].font = Font(name=FONT, bold=True, size=12)
ws.append(["Field", "Type", "Purpose"])
style_header_row(ws, 7, 3)
schema = [
    ["example_id", "text", "Unique identifier, stable across dataset versions"],
    ["domain_label", "categorical", "Ground-truth domain this example belongs to -- the profiler/gate's training target"],
    ["text", "text", "The actual prompt/example content"],
    ["is_boundary_example", "boolean", "TRUE if deliberately cross-domain (the Printer Prototype sheet shows why this matters: a calibration set with none of these produces a degenerate, meaningless threshold)"],
    ["cross_domain_hint", "text", "If boundary, which other domain(s) it also resembles -- for human review, not fed to the classifier directly"],
    ["split", "categorical", "train / calibration / test -- calibration MUST be held out separately from train (TECHNICAL.md 4.3); reusing train data for calibration is what produced the earlier degenerate-threshold failure"],
    ["source", "categorical", "hand_written / systematic_generation / real_user_log / benchmark_derived -- track provenance so calibration quality can be audited later"],
    ["added_for_version", "text", "Which profiler/dimension version introduced this example -- supports the modular composition design (TECHNICAL.md section 9)"],
    ["contributor", "text", "Who added it -- relevant if this dataset is eventually crowdsourced (e.g. via an Openfield-style marketplace) rather than built solo"],
]
r = 8
for row_ in schema:
    for c, val in enumerate(row_, start=1):
        ws.cell(row=r, column=c, value=val).font = BODY_FONT
    r += 1

# --- Sample draft rows ---
sample_header_row = r + 2
ws.cell(row=sample_header_row - 1, column=1, value="SAMPLE ENTRIES (draft)").font = Font(name=FONT, bold=True, size=12)
headers9 = ["example_id", "domain_label", "text", "is_boundary_example", "cross_domain_hint",
            "split", "source", "added_for_version", "contributor"]
for c, h in enumerate(headers9, start=1):
    ws.cell(row=sample_header_row, column=c, value=h)
style_header_row(ws, sample_header_row, 9)

samples = [
    ["ex_0001", "code", "the recursive method has a bug where the recursion never terminates", False, "", "train", "hand_written", "v1", "washington"],
    ["ex_0002", "code", "explain how a hash map resolves collisions using open addressing", False, "", "train", "hand_written", "v1", "washington"],
    ["ex_0003", "math", "prove that the series diverges using the integral test", False, "", "train", "hand_written", "v1", "washington"],
    ["ex_0004", "math", "calculate the eigenvalues of this 3x3 matrix", False, "", "calibration", "hand_written", "v1", "washington"],
    ["ex_0005", "creative", "write a scene where the protagonist realizes the narrator is unreliable", False, "", "train", "hand_written", "v1", "washington"],
    ["ex_0006", "reasoning", "identify the logical fallacy in this argument about correlation and causation", False, "", "train", "hand_written", "v1", "washington"],
    ["ex_0007", "law", "explain what makes a contract clause unconscionable under common law", False, "", "train", "hand_written", "v1", "washington"],
    ["ex_0008", "law", "the plaintiff's claim was dismissed because the statute of limitations had expired", False, "", "calibration", "hand_written", "v1", "washington"],
    ["ex_0009", "code", "review the open source license agreement attached to this repository for compliance issues", True, "law", "test", "systematic_generation", "v1", "auto"],
    ["ex_0010", "math", "calculate the statute of limitations deadline given the filing date and tolling rules", True, "law", "test", "systematic_generation", "v1", "auto"],
    ["ex_0011", "creative", "write a closing argument in the voice of a defense attorney pleading for leniency", True, "law", "test", "systematic_generation", "v1", "auto"],
    ["ex_0012", "reasoning", "assess if the contract's ambiguous clause creates a valid basis for rescission", True, "law", "test", "systematic_generation", "v1", "auto"],
    ["ex_0013", "medicine", "explain the mechanism of action for ACE inhibitors in treating hypertension", False, "", "train", "hand_written", "v2", "contributor_a"],
    ["ex_0014", "medicine", "differentiate between type 1 and type 2 diabetes presentation", False, "", "calibration", "hand_written", "v2", "contributor_a"],
    ["ex_0015", "medicine", "assess whether this AI diagnostic recommendation carries legal liability for the prescribing physician", True, "law", "test", "systematic_generation", "v2", "auto"],
    ["ex_0016", "medicine", "a patient presents with persistent cough and unexplained weight loss -- outline a differential", True, "reasoning", "test", "hand_written", "v2", "contributor_a"],
    ["ex_0017", "finance", "calculate the net present value of this cash flow series at a 7% discount rate", False, "", "train", "hand_written", "v2", "contributor_b"],
    ["ex_0018", "finance", "explain the difference between a call option and a put option", False, "", "calibration", "hand_written", "v2", "contributor_b"],
    ["ex_0019", "finance", "derive the amortization schedule required under the loan agreement's payment terms", True, "law, math", "test", "systematic_generation", "v2", "auto"],
    ["ex_0020", "finance", "prove that the bond pricing formula converges as maturity approaches zero", True, "math", "test", "systematic_generation", "v2", "auto"],
]
r = sample_header_row + 1
for row_ in samples:
    for c, val in enumerate(row_, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = BODY_FONT
        if c == 3:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c == 4 and val is True:
            cell.font = Font(name=FONT, size=10, bold=True, color="9C0006")
    r += 1
sample_end = r - 1

# --- Sizing guidance, derived from findings elsewhere in this workbook ---
guidance_row = r + 2
ws.cell(row=guidance_row, column=1, value="SIZING GUIDANCE (derived from measured findings, not guesses)").font = Font(name=FONT, bold=True, size=12)
guidance = [
    ["Examples in this draft, by domain", f"=COUNTIF(B{sample_header_row+1}:B{sample_end},\"code\")+COUNTIF(B{sample_header_row+1}:B{sample_end},\"math\")+COUNTIF(B{sample_header_row+1}:B{sample_end},\"creative\")+COUNTIF(B{sample_header_row+1}:B{sample_end},\"reasoning\")+COUNTIF(B{sample_header_row+1}:B{sample_end},\"law\")+COUNTIF(B{sample_header_row+1}:B{sample_end},\"medicine\")+COUNTIF(B{sample_header_row+1}:B{sample_end},\"finance\")"],
    ["Boundary examples in this draft", f"=COUNTIF(D{sample_header_row+1}:D{sample_end},TRUE)"],
    ["Boundary ratio in this draft", f"=B{guidance_row+2}/B{guidance_row+1}"],
    ["Recommended minimum calibration points (per domain, to avoid a degenerate threshold)", "8000 -- see Printer Prototype / Multi-Dimension Compounding sheets: this project's own tests used this scale to get a stable threshold"],
    ["Recommended boundary-example share of calibration set", "~33% -- the Printer Prototype fix used ~600 boundary examples against ~1200 clean (600/1800); a clean-only set of the same total size produced a degenerate threshold"],
    ["When adding M domains simultaneously", "Budget calibration size per TECHNICAL.md section 7's Bonferroni correction -- per-domain false-positive tolerance must shrink as M grows, or aggregate risk compounds silently"],
    ["Calibration split discipline", "Never reuse train-split examples for calibration -- this project's original text-validation failure (degenerate 0.01 threshold) came directly from calibrating on the training set instead of a genuinely held-out one"],
]
r = guidance_row + 1
for label, val in guidance:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    cell = ws.cell(row=r, column=2, value=val)
    cell.font = BODY_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if label == "Boundary ratio in this draft":
        cell.number_format = "0.0%"
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    r += 1

note_row = r + 1
ws.cell(row=note_row, column=1,
        value="This is a design draft (20 illustrative rows), not a production dataset. The schema and sizing guidance are what should scale -- "
              "the row count should not be read as a target.").font = NOTE_FONT
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)

autosize(ws, [12, 12, 46, 10, 14, 12, 18, 14, 14])
ws.freeze_panes = f"A{sample_header_row+1}"

import os
_out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "performance_data.xlsx")
wb.save(_out_path)
print("Saved performance_data.xlsx with", len(wb.sheetnames), "sheets:", wb.sheetnames)
