"""
build_lora_workbook.py

Builds lorouter_results.xlsx -- the canonical results workbook for the
lorouter branch. Numbers are this branch's canonical run (the actual
outputs of the experiments listed under 'source' below, run 2026-08-05 on
the lorouter branch). The scripts remain the source of truth; this
workbook is the summary artifact, consistent with how performance_data.xlsx
is produced on the v2 branch.

Sheets:
  Summary          -- every experiment, one line, key numbers + source
  StandinBenchmark -- 4-way strategy comparison (5 seeds, brick 2)
  RealLoRA         -- loss matrix + routing (SmolLM2-135M, 10 epochs)
  UnseenTask       -- arm-by-arm oracle agreement + loss gap (answer-NLL)
  EightAdapter     -- 8-adapter pool: accuracy, separation, isolation
  Corpus           -- brick growth + calibration thresholds (brick 2 vs 3)

Charts: one per sheet (native openpyxl charts, ink/amber palette).

Run: python3 experiments/build_lora_workbook.py
Requires: openpyxl
"""
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

OUT = Path(__file__).resolve().parent / "lorouter_results.xlsx"

INK = "0B0C0E"
CREAM = "F5EFE6"
AMBER = "FFB454"
TEAL = "2DD4BF"
RED = "F87171"

wb = openpyxl.Workbook()

HDR_FILL = PatternFill("solid", fgColor=AMBER)
HDR_FONT = Font(color=INK, bold=True, size=11)
TITLE_FONT = Font(color=CREAM, bold=True, size=14)
SUB_FONT = Font(color="A89F8D", italic=True, size=9)


def sheet(name, title, subtitle, headers, rows):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for r, row in enumerate(rows, 5):
        for c, v in enumerate(row, 1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A5"
    return ws


def bar(ws, title, cat_ref, data_refs, labels, colors, anchor="H4"):
    ch = BarChart()
    ch.type = "col"
    ch.title = title
    ch.style = 10
    for ref, lab, col in zip(data_refs, labels, colors):
        ch.add_data(ref, titles_from_data=False)
        s = ch.series[-1]
        s.tx = openpyxl.chart.series.SeriesLabel(v=lab)
        s.graphicalProperties.solidFill = col
    ch.set_categories(cat_ref)
    ch.height = 9
    ch.width = 20
    ch.gapWidth = 80
    ws.add_chart(ch, anchor)
    return ch


# ----------------------------------------------------------------------
# Summary
# ----------------------------------------------------------------------
summary_rows = [
    ["Stand-in benchmark (brick 2, 5 seeds, 56 test/seed)", "profile 96.4% | centroid 97.9% | learned 96.4% | random 19.3%", "experiments/benchmark.py", "pattern exact; digits stable across 5 seeds"],
    ["Swap isolation (text setting, stand-in)", "0 flips on other domains (weak-replacement swap)", "experiments/benchmark.py", "exact"],
    ["Real-LoRA integration (SmolLM2-135M, rank 8, 10 ep)", "routing 96.4% (54/56); differentiation 2/4 diagonal; finance/law adapters favor code (base priors)", "experiments/real_lora_integration.py", "routing pattern exact across runs; loss digits vary run-to-run (GPU nondeterminism)"],
    ["Unseen-task (education held out, 14 queries)", "routing stable: education->finance 10/14 (real + stand-in arms agree); quality verified at the aligned-adapter level (F13/F31); without aligned adapter the loss gap == random", "experiments/unseen_task_generalization.py", "distribution exact; quality numbers pattern-level"],
    ["LORAUTER-style exemplar signal (answer-NLL oracle)", "V1 42.9% / +0.13% | V2a 28.6% / +0.12% (all->code, lexical artifact) | V2b 35.7% / +0.12% | V3 aligned 14/14 routed, +0.04% | ceiling 100%", "experiments/lora_exemplar_routing.py", "pattern exact; loss digits vary"],
    ["8-adapter pool (2/domain, disjoint data)", "domain accuracy 96.4% (unchanged); profile min cosine 0.996 (near-duplicate adapters); swap flips 0/56", "experiments/eight_adapter_space.py", "pattern exact"],
    ["Corpus brick 3 v3 (user-built)", "3,010 ex / 6 domains / 220 boundaries; boundary hardness 94.0% clean vs 0.0% boundary; p99 clean-only threshold 0.0448 -> 0.7638, false-capture 1.79% -> 0.00%", "corpus/moat_brick3.jsonl + scripts/moat_calibration_trial.py", "exact (seeded/deterministic parts); v3 file fixed"],
    ["Profile-metric design (F9)", "question-only loss profiles -> 73.2% routing (code 0%); answer-conditional -> 96.4%", "experiments/real_lora_integration.py", "pattern exact"],
]
ws = sheet("Summary", "LOROUTER -- canonical results", "Canonical runs 2026-08-05/06, lorouter branch. Scripts are the source of truth.",
           ["Experiment", "Key result", "Source script", "Reproducibility"], summary_rows)
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 92
ws.column_dimensions["C"].width = 42
ws.column_dimensions["D"].width = 52

# ----------------------------------------------------------------------
# Stand-in benchmark
# ----------------------------------------------------------------------
ws = sheet("StandinBenchmark", "Stand-in benchmark -- adapter selection accuracy",
           "Brick 2 corpus machinery, 5 seeds, 56 test inputs/seed. Random floor 19.3%.",
           ["strategy", "acc seed1", "seed2", "seed3", "seed4", "seed5", "mean"],
           [
               ["profile (lorouter)", 0.964, 0.964, 0.964, 0.964, 0.964, 0.964],
               ["centroid (LORAUTER-style)", 0.982, 0.982, 0.982, 0.982, 0.964, 0.979],
               ["learned router (MoLE line)", 0.964, 0.964, 0.964, 0.964, 0.964, 0.964],
               ["random", 0.161, 0.214, 0.161, 0.214, 0.214, 0.193],
           ])
bar(ws, "Adapter selection accuracy by strategy",
    Reference(ws, min_col=1, min_row=5, max_row=8),
    [Reference(ws, min_col=7, min_row=5, max_row=8)],
    ["mean accuracy"], [AMBER])
ws["G5"].number_format = "0.0%"
ws["G6"].number_format = "0.0%"
ws["G7"].number_format = "0.0%"
ws["G8"].number_format = "0.0%"

# ----------------------------------------------------------------------
# Real LoRA
# ----------------------------------------------------------------------
ws = sheet("RealLoRA", "Real-LoRA integration -- calibration loss matrix",
           "SmolLM2-135M-Instruct, rank 8, 10 epochs, QA from brick 3. Loss = answer-conditional NLL on calibration (question masked). Lower = better.",
           ["adapter \\ calib domain", "code", "education", "finance", "law"],
           [
               ["code", 4.047, 5.259, 6.046, 6.090],
               ["education", 3.862, 3.749, 4.999, 4.835],
               ["finance", 3.892, 4.327, 4.670, 5.163],
               ["law", 4.308, 4.973, 5.850, 5.020],
           ])
ws["A9"] = "routing accuracy (profile router): 96.4% (54/56) | diagonal dominance: 2/4 | source: experiments/real_lora_integration.py"
ws["A9"].font = SUB_FONT
bar(ws, "Calibration loss by adapter and domain",
    Reference(ws, min_col=1, min_row=5, max_row=8),
    [Reference(ws, min_col=c, min_row=5, max_row=8) for c in (2, 3, 4, 5)],
    ["code", "education", "finance", "law"],
    [TEAL, AMBER, RED, "A89F8D"])
for r in range(5, 9):
    for c in range(2, 6):
        ws.cell(row=r, column=c).number_format = "0.000"

# ----------------------------------------------------------------------
# Unseen task
# ----------------------------------------------------------------------
ws = sheet("UnseenTask", "Unseen-task generalization -- arms compared",
           "Education fully held out, 14 test queries. Oracle = per-query lowest-answer-NLL adapter. Loss gap vs oracle; random expectation +0.13%.",
           ["arm", "oracle agreement", "loss gap vs oracle", "routing distribution"],
           [
               ["V1 per-query profile routing", 0.429, 0.0013, "finance 10, code 2, law 2"],
               ["V2a task-level, exemplars=5", 0.286, 0.0012, "code 14 (lexical artifact)"],
               ["V2a task-level, exemplars=10", 0.286, 0.0012, "code 14"],
               ["V2b per-query embeddings, ex=5", 0.357, 0.0012, "finance 6, code 4, law 4"],
               ["V2b per-query embeddings, ex=10", 0.357, 0.0012, "finance 6, code 4, law 4"],
               ["V3 aligned-adapter control", 0.000, 0.0004, "education 14/14 (routing perfect)"],
               ["oracle-best (ceiling)", 1.000, 0.0000, "finance 5, law 5, code 4"],
           ])
for r in range(5, 12):
    ws.cell(row=r, column=2).number_format = "0.0%"
    ws.cell(row=r, column=3).number_format = "0.00%"
bar(ws, "Oracle agreement by routing arm",
    Reference(ws, min_col=1, min_row=5, max_row=11),
    [Reference(ws, min_col=2, min_row=5, max_row=11)],
    ["oracle agreement"], [AMBER])

# ----------------------------------------------------------------------
# Calibration trial (brick 2 vs brick 3, full table)
# ----------------------------------------------------------------------
ws = sheet("CalibrationTrial", "Gate calibration -- brick 2 vs brick 3",
           "scripts/moat_calibration_trial.py, gate for education, TF-IDF+SVD+MLP. The s8 property: boundary-inclusive calibration raises the threshold and kills false-capture.",
           ["metric", "brick 2 (n_cal 42)", "v3 brick 3 (n_cal ~135/domain)", "s8 canonical"],
           [
               ["p99 clean-only threshold", 0.3048, 0.0448, 0.0031],
               ["p99 clean+boundary threshold", 0.9920, 0.7638, 0.9943],
               ["p99 false-capture clean-only", 0.0476, 0.0179, 0.0160],
               ["p99 false-capture clean+boundary", 0.0000, 0.0000, 0.0000],
               ["p99 recall clean-only", 1.000, 0.973, 1.000],
               ["p99 recall clean+boundary", 0.714, 0.933, 0.920],
               ["p95 clean-only threshold", 0.1361, 0.0032, None],
               ["p95 clean+boundary threshold", 0.9353, 0.0594, None],
               ["p95 false-capture clean-only", 0.0476, 0.0402, None],
               ["p95 false-capture clean+boundary", 0.0000, 0.0134, None],
               ["p95 recall clean-only", 1.000, 0.987, None],
               ["p95 recall clean+boundary", 0.929, 0.973, None],
               ["addition flips (joint retrain, part 1)", 2, 7, "see s6"],
               ["education-boundary escalation (p99)", "1 of 7", "6 of 34", None],
           ])
for r in range(5, 19):
    for c in (2, 3, 4):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, float):
            ws.cell(row=r, column=c).number_format = "0.0000"
bar(ws, "p95 thresholds: brick 2 vs brick 3",
    Reference(ws, min_col=1, min_row=5, max_row=12),
    [Reference(ws, min_col=2, min_row=5, max_row=12),
     Reference(ws, min_col=3, min_row=5, max_row=12)],
    ["brick 2", "brick 3"], [RED, TEAL], anchor="G4")

# ----------------------------------------------------------------------
# Profile metric design (F9)
# ----------------------------------------------------------------------
ws = sheet("ProfileMetric", "Profile metric design (finding F9)",
           "experiments/real_lora_integration.py, same adapters, two profile definitions. Profiles must measure ANSWER behavior, not question text.",
           ["profile metric", "routing accuracy", "code domain", "note"],
           [
               ["question-only NLL (early design)", 0.732, 0.0, "near-uniform profiles after normalization; code collapsed"],
               ["answer-conditional NLL (current)", 0.964, 0.929, "signal restored; matches stand-in benchmark"],
           ])
for r in range(5, 7):
    ws.cell(row=r, column=2).number_format = "0.0%"
    ws.cell(row=r, column=3).number_format = "0.0%"
bar(ws, "Routing accuracy by profile metric",
    Reference(ws, min_col=1, min_row=5, max_row=6),
    [Reference(ws, min_col=2, min_row=5, max_row=6)],
    ["accuracy"], [AMBER], anchor="E4")

# ----------------------------------------------------------------------
# Adapter training (final losses, canonical runs)
# ----------------------------------------------------------------------
ws = sheet("AdapterTraining", "Adapter training outcomes (final losses)",
           "SmolLM2-135M-Instruct, rank 8, 10 epochs, 62 QA pairs per domain (4-adapter runs; 31 per variant in the 8-adapter run). GPU nondeterminism: digits vary run-to-run, patterns stable.",
           ["adapter", "final train loss (4-adapter run)", "final train loss (8-adapter run)", "note"],
           [
               ["code", 1.368, None, "code-prior base model favors this domain for all adapters"],
               ["education", 1.473, None, "steepest descent; overfits its own QA style"],
               ["finance", 1.555, None, ""],
               ["law", 1.389, None, ""],
               ["code_A / code_B", None, "1.477 / (B)", "8-adapter variants trained on 31 QA pairs"],
               ["law_A / law_B", None, "(A) / 1.582", ""],
           ])

# ----------------------------------------------------------------------
# Gap-closure sheets (2026-08-06)
# ----------------------------------------------------------------------
ws = sheet("SemanticBenchmark", "Semantic embeddings arm (bge-small-en-v1.5)",
           "experiments/semantic_embeddings.py, brick 2, 5 seeds (same corpus as the TF-IDF benchmark -- like-for-like). Semantic features improve profile routing to 98.2% and flip the centroid ranking (F28).",
           ["strategy", "semantic acc", "TF-IDF+SVD acc", "note"],
           [
               ["profile (lorouter)", 0.9821, 0.9643, "now BEATS centroid; ties learned"],
               ["centroid (embedding)", 0.9643, 0.9786, "ranking flipped vs lexical features"],
               ["learned router", 0.9821, 0.9643, ""],
               ["random", 0.1929, 0.1929, "floor"],
           ])
for r in range(5, 9):
    for c in (2, 3):
        ws.cell(row=r, column=c).number_format = "0.00%"
bar(ws, "Semantic vs TF-IDF features: profile routing",
    Reference(ws, min_col=1, min_row=5, max_row=8),
    [Reference(ws, min_col=2, min_row=5, max_row=8),
     Reference(ws, min_col=3, min_row=5, max_row=8)],
    ["semantic", "TF-IDF+SVD"], [TEAL, RED], anchor="F4")
ws = sheet("UnseenSemantic", "Unseen-task exemplar routing: F12 retest",
           "experiments/semantic_embeddings.py. The lexical artifact (all->code) disappears with semantic task embeddings (F29).",
           ["features", "task-level (10 exemplars)", "per-query distribution"],
           [
               ["TF-IDF+SVD (old)", "code (artifact)", "finance 10, code 2, law 2"],
               ["bge-small-en-v1.5", "finance (sensible)", "finance 6, code 3, law 5"],
           ])
ws = sheet("ModelScale", "Model-size + seed invariance",
           "real_lora_multiseed.py (seeds 42/7/2026) + real_lora_360m.py. Routing accuracy is stable across seeds and model sizes (F26, F27).",
           ["model", "seeds", "routing acc", "diagonal dominance"],
           [
               ["SmolLM2-135M", "42, 7, 2026", "96.4 / 96.4 / 96.4", "1/4 each"],
               ["SmolLM2-360M", "42", "96.4%", "2/4"],
           ])
ws = sheet("Latency", "Selection-policy latency spike",
           "experiments/latency_spike.py, warm mean of 1000 queries. Policy is sub-ms at any realistic pool size (F30).",
           ["pool size N", "cosine selection (us)", "query profiling (ms)", "end-to-end (ms)"],
           [
               [4, 16.9, 1.140, 1.157],
               [8, 26.6, 1.140, 1.167],
               [100, 24.5, 1.140, 1.165],
               [1000, 65.2, 1.140, 1.205],
               [10000, 379.5, 1.140, 2.775],
           ])
for r in range(5, 10):
    ws.cell(row=r, column=2).number_format = "0.0"
lc = LineChart()
lc.title = "Selection latency vs pool size"
lc.style = 12
lc.height = 9
lc.width = 20
data = Reference(ws, min_col=2, min_row=4, max_row=9)
cats = Reference(ws, min_col=1, min_row=5, max_row=9)
lc.add_data(data, titles_from_data=True)
lc.set_categories(cats)
ws.add_chart(lc, "F4")
ws = sheet("GenerationQuality", "Aligned-adapter generation quality",
           "experiments/generation_quality.py: education held out, generations scored by bge-small similarity to reference (F31).",
           ["adapter", "mean similarity", "note"],
           [
               ["education (aligned)", 0.5932, "routed choice; best output; only domain-plausible continuation in samples"],
               ["code (best incumbent)", 0.5887, ""],
               ["law", 0.5839, "random expectation 0.5862"],
               ["finance", 0.5792, ""],
           ])
for r in range(5, 9):
    ws.cell(row=r, column=2).number_format = "0.0000"
bar(ws, "Generation similarity to reference by adapter",
    Reference(ws, min_col=1, min_row=5, max_row=8),
    [Reference(ws, min_col=2, min_row=5, max_row=8)],
    ["similarity"], [AMBER], anchor="D4")

# ----------------------------------------------------------------------
# Adapter-pool scaling (F32-F36)
# ----------------------------------------------------------------------
ws = sheet("PoolScaling", "Adapter-pool scaling (F32-F36)",
           "experiments/adapter_pool_scaling.py: variant profiles = real-LoRA base profiles + Gaussian noise sigma, 5 seeds, 92 brick-3 queries. U-shaped accuracy: sigma=0 flat at 96.7%; noise -> mid-N helps, extreme-N decays.",
           ["N", "sigma", "accuracy", "false_capture", "sep_min", "sep_mean"],
           [
               [8, 0.00, 0.9674, 0.0326, 0.993, 0.996],
               [32, 0.00, 0.9674, 0.0326, 0.993, 0.997],
               [128, 0.00, 0.9674, 0.0326, 0.993, 0.997],
               [512, 0.00, 0.9674, 0.0326, 0.993, 0.997],
               [1024, 0.00, 0.9674, 0.0326, 0.993, 0.997],
               [8, 0.05, 0.3957, 0.6043, 0.902, 0.964],
               [32, 0.05, 0.5978, 0.4022, 0.846, 0.968],
               [128, 0.05, 0.7739, 0.2261, 0.778, 0.968],
               [512, 0.05, 0.7304, 0.2696, 0.743, 0.968],
               [1024, 0.05, 0.6565, 0.3435, 0.715, 0.968],
               [8, 0.10, 0.3870, 0.6130, 0.688, 0.883],
               [32, 0.10, 0.4391, 0.5609, 0.431, 0.886],
               [128, 0.10, 0.6630, 0.3370, 0.283, 0.887],
               [512, 0.10, 0.4391, 0.5609, 0.021, 0.887],
               [1024, 0.10, 0.4761, 0.5239, -0.065, 0.885],
               [8, 0.20, 0.3717, 0.6283, 0.061, 0.641],
               [32, 0.20, 0.3804, 0.6196, -0.479, 0.639],
               [128, 0.20, 0.3891, 0.6109, -0.827, 0.642],
               [512, 0.20, 0.2130, 0.7870, -0.982, 0.635],
               [1024, 0.20, 0.3043, 0.6957, -0.997, 0.630],
           ])
for r in range(5, 25):
    ws.cell(row=r, column=3).number_format = "0.00%"
    ws.cell(row=r, column=4).number_format = "0.00%"
lc = LineChart()
lc.title = "Routing accuracy vs pool size (by sigma)"
lc.style = 12
lc.height = 9
lc.width = 22
for i, sig in enumerate([0.00, 0.05, 0.10, 0.20]):
    start = 5 + i * 5
    data = Reference(ws, min_col=3, min_row=start, max_row=start + 4)
    lc.add_data(data, titles_from_data=False)
    lc.series[-1].tx = openpyxl.chart.series.SeriesLabel(v=f"sigma={sig}")
cats = Reference(ws, min_col=1, min_row=5, max_row=9)
lc.set_categories(cats)
ws.add_chart(lc, "H4")

# ----------------------------------------------------------------------
# Eight adapter
# ----------------------------------------------------------------------
ws = sheet("EightAdapter", "Eight-adapter pool (2/domain)",
           "Disjoint training data per variant, different answer wording. Domain-level accuracy is the meaningful metric (both variants correct).",
           ["metric", "value", "note"],
           [
               ["domain-level routing accuracy", 0.964, "54/56; identical to 4-adapter pool"],
               ["profile min pairwise cosine", 0.996, "near-duplicate adapters -> near-identical profiles"],
               ["profile mean pairwise cosine", 0.998, "separation bounded by adapter diversity"],
               ["swap isolation (code_A swapped)", 0, "0/56 routing flips on other domains"],
               ["variant split (routed, info only)", "code 0A/13B, edu 16A/0B, fin 0A/14B, law 13A/0B", "router consistently prefers one variant per domain"],
           ])
ws["A10"] = "source: experiments/eight_adapter_space.py"
ws["A10"].font = SUB_FONT
bar(ws, "Routing accuracy: 4-adapter vs 8-adapter pool",
    Reference(ws, min_col=1, min_row=5, max_row=6),
    [Reference(ws, min_col=2, min_row=5, max_row=6)],
    ["accuracy"], [TEAL])
ws["B5"].number_format = "0.0%"

# ----------------------------------------------------------------------
# Corpus growth
# ----------------------------------------------------------------------
ws = sheet("Corpus", "Moat corpus growth + calibration stability",
           "Bricks built by scripts/build_moat_corpus.py (seeded, deterministic). Thresholds from scripts/moat_calibration_trial.py (gate: education).",
           ["brick", "total examples", "boundaries", "calibration split", "p99 clean-only thr", "p99 false-capture", "p95 clean-only thr", "p95 clean+boundary thr"],
           [
               ["brick 1", 181, 17, "70/15/15 (n~32)", None, None, None, None],
               ["brick 2", 408, 48, "70/15/15 (n=80)", 0.3048, 0.0476, 0.1361, 0.9353],
               ["brick 3 (pre-v3)", 664, 64, "60/25/15 (n=180)", 0.8439, 0.0000, 0.1283, 0.6029],
               ["brick 3 v3 (user-built)", 3010, 220, "1673/807/530 (n~135/domain)", 0.0448, 0.0179, 0.0032, 0.0594],
           ])
ws["A9"] = "brick 2 -> pre-v3: small-n p99 degeneracy resolved. v3 (user-built, 6 domains + medicine/psychology): the six-domain space is harder -- p99 false-capture 1.79% -> 0.00% with boundaries; p95 effect real but weaker (4.02% -> 1.34%). Addition flips at v3: 7."
ws["A9"].font = SUB_FONT
ws["A10"] = "v3 integrity: 3,010 unique ids/texts, zero calib/test leakage into train; boundary hardness 94.0% clean vs 0.0% boundary (F37-F41)."
ws["A10"].font = SUB_FONT
lc = LineChart()
lc.title = "Corpus growth across bricks"
lc.style = 12
lc.height = 9
lc.width = 20
data = Reference(ws, min_col=2, min_row=4, max_row=7)
cats = Reference(ws, min_col=1, min_row=5, max_row=7)
lc.add_data(data, titles_from_data=True)
lc.set_categories(cats)
ws.add_chart(lc, "H4")
bar(ws, "Calibration thresholds: brick 2 vs brick 3 (p95)",
    Reference(ws, min_col=1, min_row=5, max_row=6),
    [Reference(ws, min_col=7, min_row=5, max_row=6),
     Reference(ws, min_col=8, min_row=5, max_row=6)],
    ["clean-only", "clean+boundary"], [RED, TEAL], anchor="H23")

# drop the default sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(OUT)
print(f"wrote {OUT}")
print("sheets:", wb.sheetnames)
